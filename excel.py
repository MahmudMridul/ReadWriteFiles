import openpyxl as excel

file_path = 'file.xlsx'

def get_sheetnames(filepath):
    workbook = excel.load_workbook(filepath)
    print(workbook.sheetnames)


def show_sheet(filepath, sheetname):
    workbook = excel.load_workbook(filepath)
    sheet = workbook[sheetname]

    rows = sheet.max_row
    cols = sheet.max_column

    for row in range(1, rows+1):
        for col in range(1, cols+1):
            cell = sheet.cell(row = row, column = col)
            print(cell.value, end=' ')
        print()

def create_newsheet(filepath, sheetname):
    workbook = excel.load_workbook(filepath)
    workbook.create_sheet(sheetname)
    workbook.save(filepath)

def create_newworkbook(filepath):
    workbook = excel.Workbook()
    workbook.save(filepath)


